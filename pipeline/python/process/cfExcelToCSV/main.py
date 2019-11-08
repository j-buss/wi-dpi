from google.cloud import storage
import openpyxl
import pandas as pd
import os


def excel_to_csv(data, context):
    # Config
    source_bucket_name = data['bucket']
    source_blob_name = data['name']
    source_blob_basename, source_blob_ext = os.path.splitext(os.path.basename(source_blob_name))

    if source_blob_ext == ".xlsx":
        print("Converting workbook: " + source_blob_name)
        # Get source blob in usable state
        storage_client = storage.Client()
        source_bucket = storage_client.get_bucket(source_bucket_name)
        source_blob = source_bucket.blob(source_blob_name)

        data_xls = pd.read_excel(source_blob, 'Sheet1', index_col=None)
        data_xls.to_csv('your_csv.csv', encoding='utf-8')

        wb = openpyxl.load_workbook(source_blob)
        # wb = openpyxl.load_workbook("gs://landing-008/Test_Spreadsheet.xlsx")
        # Get source blob in usable state

        # Cycle through spreadsheet - convert one tab at a time
        for sheet in wb.sheetnames:

            # For each Sheet
            print("Converting sheet: " + sheet)
            csv_file_name = source_blob_basename + "_" + sheet + ".csv"
            print("Creating file: " + csv_file_name)
            source_sheet = wb.get_sheet_by_name(sheet)
            sheet_max_col = source_sheet.max_column
            sheet_max_row = source_sheet.max_row

            # Create Target File
            # target_file = open(csv_file_name, 'w')
            # target_writer = csv.writer(target_file)

            # for row in range(1, sheet_max_row + 1):
            #     rowArray = []
            # for column in range(1, sheet_max_col):
            # rowArray.append(source_sheet.cell(row=row, column=column).value)

    # target_writer.writerow(rowArray)
    # target_blob = target_bucket.blob(contentfilename)
    # target_blob.upload_from_string(contentfile)
# targetFile.close()